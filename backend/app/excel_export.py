import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
import logging

logger = logging.getLogger(__name__)


def generate_excel_report(events: list, stats: dict, filename: str = None) -> str:
    """Generate Excel report with formatting"""
    if not filename:
        filename = f"reports/attack_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    try:
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Attack Events"

        # Add title
        ws['A1'] = "HoneyCloud-X Attack Report"
        ws['A1'].font = Font(size=16, bold=True, color="00FFFFFF")
        ws['A1'].fill = PatternFill(start_color="0A0E27", end_color="0A0E27", fill_type="solid")
        ws.merge_cells('A1:I1')

        # Add timestamp
        ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A2'].font = Font(size=10, italic=True)
        ws.merge_cells('A2:I2')

        # Add statistics section
        row = 4
        ws[f'A{row}'] = "STATISTICS"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        
        row += 1
        ws[f'A{row}'] = "Total Events"
        ws[f'B{row}'] = stats.get('total_events', 0)
        ws[f'B{row}'].font = Font(bold=True, color="00FF6B6B")

        row += 1
        ws[f'A{row}'] = "Critical Events"
        ws[f'B{row}'] = stats.get('events_by_severity', {}).get('CRITICAL', 0)
        ws[f'B{row}'].font = Font(bold=True, color="00FF0000")

        row += 1
        ws[f'A{row}'] = "Malicious (AI)"
        ws[f'B{row}'] = stats.get('ai_labels', {}).get('malicious', 0)
        ws[f'B{row}'].font = Font(bold=True, color="00FF9500")

        # Add events table
        row = 12
        headers = ['ID', 'Timestamp', 'Service', 'Source IP', 'Username', 'Severity', 'AI Label', 'Score', 'Command']
        
        # Header row
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="00FFFFFF")
            cell.fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        row += 1
        for event in events:
            ws.cell(row=row, column=1).value = event.get('id')
            ws.cell(row=row, column=2).value = event.get('timestamp')
            ws.cell(row=row, column=3).value = event.get('service')
            ws.cell(row=row, column=4).value = event.get('source_ip')
            ws.cell(row=row, column=5).value = event.get('username')
            
            severity = event.get('severity')
            severity_cell = ws.cell(row=row, column=6)
            severity_cell.value = severity
            if severity == 'CRITICAL':
                severity_cell.fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
            elif severity == 'HIGH':
                severity_cell.fill = PatternFill(start_color="FFFF9500", end_color="FFFF9500", fill_type="solid")
            
            ws.cell(row=row, column=7).value = event.get('ai_label')
            ws.cell(row=row, column=8).value = event.get('threat_score')
            ws.cell(row=row, column=9).value = event.get('command')
            
            row += 1

        # Adjust column widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 10
        ws.column_dimensions['I'].width = 25

        # Save file
        wb.save(filename)
        logger.info(f"✅ Excel report generated: {filename}")
        return filename

    except Exception as e:
        logger.error(f"❌ Error generating Excel: {e}")
        raise
